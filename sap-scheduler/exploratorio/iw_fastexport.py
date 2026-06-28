"""Exportação RÁPIDA (VIA C) pra IW47/IW37: navega + adiciona colunas (move-all no
layout) + dispara o export ALV nativo &XXL -> .xlsx (SAP grava TUDO num disparo, sem
paginar célula-a-célula) + kill-excel. Depois converte o .xlsx -> CSV com NOMES TÉCNICOS
(mapeados por posição via ColumnOrder) subsetado nas colunas-alvo.

Reusa o export_grid já provado no zpp_gridcap.py (ZPPPRD/NT0001). READ-ONLY no SAP
(export grava arquivo em disco; mudança de layout é só exibição, não salva variante).

Uso:
  python iw_fastexport.py --tx iw47 --low 01.03.2026 --high 27.05.2026 --out IW47_realizado_3meses.csv
  python iw_fastexport.py --tx iw37 --low 01.03.2026 --high 27.05.2026 --out IW37_planejado_3meses.csv
  (--inspect: só exporta o xlsx e mostra dims/tipos das colunas-alvo, sem gerar CSV)
"""
import time, csv, argparse, os, datetime
from types import SimpleNamespace
import win32com.client
import zpp_gridcap as z

TARGET_IW47 = ["AUFNR", "VORNR", "WERKI", "AUART", "STTXT", "PERNR", "NAME", "TPLNR",
               "ISMNW", "ISMNE", "ARBEI", "ARBEH", "PARBPL", "IARBPL",
               "ISDD", "ISDZ", "IEDD", "IEDZ", "ERSDA", "BUDAT",
               "STOKZ", "STZHL", "RUECK", "RMZHL"]
TARGET_IW37 = ["AUFNR", "VORNR", "AUART", "LTXA1", "VSTTXT", "PRIOK", "ARBPL", "IWERK", "WERKS",
               "TPLNR", "EQUNR", "EQKTX", "ARBEI", "ARBEH",
               "FSAVD", "FSAVZ", "FSEDD", "FSEDZ", "SSAVD", "SSAVZ", "SSEDD", "SSEDZ",
               "ISDD", "IEDD", "PERNR"]


def eng():
    try:
        return win32com.client.GetObject("SAPGUI").GetScriptingEngine
    except Exception:
        w = win32com.client.Dispatch("SapROTWr.SapROTWrapper")
        return w.GetROTEntry("SAPGUI").GetScriptingEngine


def open_layout_and_move_all(s):
    found = []

    def walk(node):
        try:
            kids = node.Children; n = kids.Count
        except Exception:
            return
        for i in range(n):
            try:
                ch = kids(i)
            except Exception:
                continue
            if "atuais" in (getattr(ch, "Text", "") or "").lower():
                found.append(getattr(ch, "Id", ""))
            walk(ch)
    walk(s.findById("wnd[0]/mbar"))
    if not found:
        print("   [erro] menu 'Atuais...' não achado"); return False
    s.findById(found[0]).select(); time.sleep(1.0)
    base = "wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/ssubSUB_DYN0510:SAPLSKBH:0620/"
    setgrid = s.findById(base + "cntlCONTAINER1_LAYO/shellcont/shell")
    n = setgrid.RowCount
    setgrid.selectedRows = f"0-{n-1}"
    s.findById(base + "btnAPP_WL_SING").press(); time.sleep(0.8)
    s.findById("wnd[1]/tbar[0]/btn[0]").press(); time.sleep(1.5)
    return True


def navigate(s, tx, low, high):
    s.findById("wnd[0]/tbar[0]/okcd").Text = f"/n{tx}"
    s.findById("wnd[0]").sendVKey(0); time.sleep(0.8)
    if tx == "iw47":
        for k, v in (("ctxtVARIANT", "/TRAB REAL"), ("ctxtWERKS_C-LOW", "BR02"),
                     ("ctxtERSDA_C-LOW", low), ("ctxtERSDA_C-HIGH", high)):
            try:
                s.findById("wnd[0]/usr/" + k).Text = v
            except Exception as e:
                print(f"   [aviso] {k}: {e}")
    else:  # iw37
        for chk in ("chkDY_AKT", "chkDY_HIS"):
            try:
                s.findById("wnd[0]/usr/" + chk).selected = True
            except Exception:
                pass
        for k, v in (("ctxtWERKS-LOW", "BR02"), ("ctxtFSAVD-LOW", low), ("ctxtFSAVD-HIGH", high)):
            try:
                s.findById("wnd[0]/usr/" + k).Text = v
            except Exception as e:
                print(f"   [aviso] {k}: {e}")
    s.findById("wnd[0]").sendVKey(8); time.sleep(1.8)


def fmt(v):
    """valor nativo do xlsx -> string PT-cru consistente."""
    if v is None:
        return ""
    if isinstance(v, datetime.datetime):
        if v.hour == 0 and v.minute == 0 and v.second == 0:
            return v.strftime("%d.%m.%Y")
        return v.strftime("%d.%m.%Y %H:%M:%S")
    if isinstance(v, datetime.date):
        return v.strftime("%d.%m.%Y")
    if isinstance(v, datetime.time):
        return v.strftime("%H:%M:%S")
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else repr(v)
    return str(v)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--tx", choices=["iw47", "iw37"], required=True)
    ap.add_argument("--low", default="01.03.2026")
    ap.add_argument("--high", default="27.05.2026")
    ap.add_argument("--out", required=True)
    ap.add_argument("--inspect", action="store_true")
    a = ap.parse_args()
    s = eng().Children(0).Children(0)
    print(f"[OK] {s.Info.SystemName}/{s.Info.Client} user={s.Info.User} | tx={a.tx}")
    for w in ("wnd[2]", "wnd[1]"):
        try:
            s.findById(w).sendVKey(12); time.sleep(0.3)
        except Exception:
            pass

    t0 = time.time()
    navigate(s, a.tx, a.low, a.high)
    grid = z.find_grid(s.findById("wnd[0]/usr"))
    print(f"   grid RowCount={grid.RowCount}")
    if not open_layout_and_move_all(s):
        return
    grid = z.find_grid(s.findById("wnd[0]/usr"))
    cols_tech = [str(c) for c in z.coll(grid.ColumnOrder) if c is not None]
    # título de exibição de cada coluna técnica (na ordem) — p/ alinhar com o xlsx
    disp_tech = [(c, z.col_title(grid, c)) for c in cols_tech]
    print(f"   colunas (técnicas, em ordem): {len(cols_tech)}")
    t_layout = time.time() - t0

    # EXPORT VIA C (&XXL nativo) — um disparo, pega tudo
    xlsx = os.path.splitext(a.out)[0] + "_viaC.xlsx"
    args = SimpleNamespace(tx=None, set=[], no_f8=False, ctx_code="&XXL",
                           kill_excel=True, save_dir=os.getcwd(),
                           fname=os.path.basename(xlsx))
    t1 = time.time()
    z.export_grid(s, args)
    t_export = time.time() - t1
    if not os.path.exists(xlsx):
        print(f"   [ERRO] xlsx não gerado: {xlsx}"); return

    # lê o xlsx e mapeia POSICIONALMENTE p/ nomes técnicos do ColumnOrder
    from openpyxl import load_workbook
    t2 = time.time()
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header_disp = list(next(rows))
    print(f"   xlsx: {ws.max_row} linhas x {ws.max_column} colunas (header display)")
    # ALINHAMENTO SEQUENCIAL por título de exibição: o &XXL exporta na ordem do grid,
    # mas descarta colunas de controle (PM_SELECTED/ROW_COUNT). Percorre o xlsx e avança
    # no disp_tech até casar o título -> resolve descartadas E duplicatas de título.
    tech_by_pos = [None] * len(header_disp)
    ti = 0
    dropped = []
    for xi, h in enumerate(header_disp):
        h = (h or "")
        start = ti
        while ti < len(disp_tech) and (disp_tech[ti][1] or "") != h:
            dropped.append(disp_tech[ti][0]); ti += 1
        if ti < len(disp_tech):
            tech_by_pos[xi] = disp_tech[ti][0]; ti += 1
        else:
            # título não casou em sequência -> tenta achar em qualquer posição restante
            ti = start
            print(f"   [aviso] header xlsx[{xi}]={h!r} não casou em sequência")
    print(f"   colunas descartadas pelo export (controle): {dropped}")
    target = TARGET_IW47 if a.tx == "iw47" else TARGET_IW37
    idx = {t: tech_by_pos.index(t) for t in target if t in tech_by_pos}
    miss = [t for t in target if t not in tech_by_pos]
    print(f"   alvo presentes: {sorted(idx)} | AUSENTES: {miss}")

    if a.inspect:
        print("   --- header xlsx -> técnico (alinhamento) ---")
        for xi, h in enumerate(header_disp):
            print(f"     xlsx[{xi:2}] {str(h)[:28]:28} -> {tech_by_pos[xi]}")
        print("   --- amostra (3 linhas, valor + tipo) das colunas-alvo ---")
        for ri, row in enumerate(rows):
            if ri >= 3:
                break
            print("     ", {t: (row[idx[t]], type(row[idx[t]]).__name__) for t in list(idx)[:12]})
        return

    out_cols = [t for t in target if t in idx]
    with open(a.out, "w", newline="", encoding="utf-8") as fh:
        wtr = csv.writer(fh)
        wtr.writerow(out_cols)
        cnt = 0
        for row in rows:
            wtr.writerow([fmt(row[idx[t]]) for t in out_cols])
            cnt += 1
    t_csv = time.time() - t2
    print(f"   [OK] {a.out} - {cnt} linhas + cabecalho | header: {out_cols}")
    print(f"   TEMPOS: nav+layout={t_layout:.0f}s export&XXL={t_export:.0f}s xlsx->csv={t_csv:.0f}s "
          f"TOTAL={time.time()-t0:.0f}s")


if __name__ == "__main__":
    main()
