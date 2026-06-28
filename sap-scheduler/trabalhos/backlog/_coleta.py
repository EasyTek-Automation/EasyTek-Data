# -*- coding: utf-8 -*-
"""PORT da coleta de backlog (IW39 + IW37) -> coletar_ordens(session).

Validado ao vivo na BR02 (GM3/040, 40CPMANUT) — IM-03. READ-ONLY no SAP quanto a
dados de negócio. NÃO classifica (categoria/disciplina/subclassificação são feitas em
_classifica.py).

Implementação (decisão IM-03, 2026-06-27):
- IW39 e IW37 selecionam a variante de exibição GLOBAL /BACKLOG na tela de seleção
  (campo `ctxtVARIANT`) → layout determinístico; a /BACKLOG do IW37 já traz a coluna
  ARBPL, sem mexer em coluna oculta a cada run (fallback de expor coluna só se a
  variante sumir do SAP).
- Leitura do grid por EXPORT &XXL (ALV→xlsx) → openpyxl → deleta o arquivo. Bem mais
  rápido que célula-a-célula no volume do IW37 (~16k linhas de operação): ~2,5 min vs
  ~7,6 min, sem trava.
- O &XXL auto-abre o Excel; ele é morto no fim (kill por PID novo). Depende do aviso de
  Segurança SAPGUI estar DESABILITADO no PC-cliente (premissa de ambiente já garantida).

Deps Windows-only (`win32clipboard`, `openpyxl`) são importadas LAZY dentro das funções
que as usam — assim o módulo importa em Linux (testes/mock) sem quebrar, igual ao padrão
de custos_manut/custo_collect.py.
"""
from __future__ import annotations

import os
import subprocess
import tempfile
import time
from datetime import datetime

# --- Constantes de coleta (BR-01, BR-02) -------------------------------------------------
AUART_BACKLOG = ["YPM1", "YPM2", "YPM8", "YPM9"]   # tipos que entram (exclui PM04/ZPM*)
VARIANTE = "/BACKLOG"                                # variante global IW39/IW37 (layout fixo)


def _log(m):
    print(f"[coleta] {m}", flush=True)


def _clip_set(vals):
    """Só p/ a multisseleção de AUFNR/AUART (Upload do clipboard do SAP). Win-only (lazy)."""
    import win32clipboard  # noqa: PLC0415 — Windows-only, importado sob demanda
    win32clipboard.OpenClipboard()
    win32clipboard.EmptyClipboard()
    win32clipboard.SetClipboardText("\r\n".join(vals))
    win32clipboard.CloseClipboard()


def _drain(s):
    for w in ("wnd[2]", "wnd[1]"):
        try:
            s.findById(w).sendVKey(12)
        except Exception:
            pass


def _coll(c):
    out = []
    try:
        n = c.Count
    except Exception:
        return out
    for i in range(n):
        v = None
        for acc in ("ElementAt", "Item"):
            try:
                v = getattr(c, acc)(i)
                break
            except Exception:
                pass
        if v is None:
            try:
                v = c(i)
            except Exception:
                v = None
        out.append(v)
    return out


def _find_grid(s, root="wnd[0]/usr"):
    def fg(elm, d=0):
        try:
            t = getattr(elm, "Type", "")
            if t == "GuiGridView":
                return elm
            if t in ("GuiShell", "GuiCtrlGridView"):
                _ = elm.ColumnOrder.Count
                return elm
        except Exception:
            pass
        if d > 45:
            return None
        try:
            k = elm.Children
            for i in range(k.Count):
                r = fg(k(i), d + 1)
                if r:
                    return r
        except Exception:
            pass
        return None

    return fg(s.findById(root))


def _excel_pids():
    pids = set()
    try:
        out = subprocess.run('tasklist /FI "IMAGENAME eq EXCEL.EXE" /FO CSV /NH',
                             shell=True, capture_output=True, text=True).stdout
        for line in out.splitlines():
            parts = [p.strip().strip('"') for p in line.split('","')]
            if len(parts) >= 2 and parts[0].upper().startswith("EXCEL"):
                pids.add(parts[1])
    except Exception:
        pass
    return pids


def _kill_new_excel(before, wait=20):
    """Espera surgir Excel novo (o &XXL auto-abre) e mata SÓ esse. Arquivo já em disco."""
    deadline = time.time() + wait
    while time.time() < deadline:
        novos = _excel_pids() - before
        if novos:
            for pid in novos:
                os.system(f"taskkill /PID {pid} /F >nul 2>&1")
            return novos
        time.sleep(0.4)
    return set()


def _export_xlsx(s, grid, target):
    """Dispara &XXL (ALV->planilha), salva em `target`, mata o Excel auto-aberto.
    Pré-requisito (garantido no ambiente do cliente): aviso de Segurança SAPGUI
    desabilitado. Devolve True se o arquivo foi criado."""
    save_dir = os.path.dirname(target)
    fname = os.path.basename(target)
    if os.path.exists(target):
        try:
            os.remove(target)
        except Exception:
            pass
    before = _excel_pids()
    try:
        grid.setCurrentCell(-1, "")
    except Exception:
        pass
    grid.contextMenu()
    grid.selectContextMenuItem("&XXL")
    time.sleep(1.0)
    # POPUP 1 — formato: escolher XLSX
    try:
        cmb = s.findById("wnd[1]/usr/cmbG_LISTBOX", False)
        if cmb is not None:
            try:
                s.findById("wnd[1]/usr/radRB_OTHERS").select()
            except Exception:
                pass
            ents = []
            try:
                es = cmb.Entries
                for i in range(es.Count):
                    e = es.ElementAt(i) if hasattr(es, "ElementAt") else es(i)
                    ents.append((e.Key, str(e.Value)))
            except Exception:
                pass
            key = next((k for k, v in ents
                        if "XLSX" in v.upper() or "OFFICE OPEN" in v.upper()), None)
            if key is not None:
                cmb.Key = key
            s.findById("wnd[1]").sendVKey(0)
            time.sleep(1.0)
    except Exception:
        pass
    # POPUP 2 — salvar (path/filename)
    try:
        s.findById("wnd[1]/usr/ctxtDY_PATH").Text = save_dir
        s.findById("wnd[1]/usr/ctxtDY_FILENAME").Text = fname
        s.findById("wnd[1]/tbar[0]/btn[0]").press()
        time.sleep(1.5)
    except Exception as ex:
        _log(f"export: popup de salvar ausente/diferente: {ex}")
    # follow-ups (Substituir / Enter)
    for _ in range(3):
        w = s.findById("wnd[1]", False)
        if w is None:
            break
        pressed = False
        try:
            s.findById("wnd[1]/tbar[0]/btn[11]").press()
            pressed = True
        except Exception:
            pass
        if not pressed:
            try:
                w.sendVKey(0)
            except Exception:
                break
        time.sleep(0.8)
    _kill_new_excel(before, wait=20)
    return os.path.exists(target)


def _export_and_read(s, grid, wanted, target):
    """Exporta o grid p/ xlsx (&XXL), lê com openpyxl mapeando colunas por TÍTULO de
    exibição (variantes via GetColumnTitles), deleta o arquivo. Devolve list[dict]
    keyed por ID técnico. Rápido mesmo com dezenas de milhares de linhas."""
    from openpyxl import load_workbook  # noqa: PLC0415 — importado sob demanda
    variants = {}
    for tech in wanted:
        vs = set()
        try:
            for t in _coll(grid.GetColumnTitles(tech)):
                if t:
                    vs.add(str(t).strip())
        except Exception:
            pass
        variants[tech] = vs
    if not _export_xlsx(s, grid, target):
        _log("export: arquivo não criado -> []")
        return []
    out = []
    try:
        wb = load_workbook(target, read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header = [(str(h).strip() if h is not None else "") for h in next(it)]
        idx = {}
        for tech in wanted:
            for j, h in enumerate(header):
                if h in variants[tech]:
                    idx[tech] = j
                    break
        for r in it:
            rec = {}
            for tech, j in idx.items():
                v = r[j] if j < len(r) else None
                rec[tech] = "" if v is None else str(v).strip()
            out.append(rec)
        wb.close()
    finally:
        try:
            os.remove(target)
        except Exception:
            pass
    return out


def _parse_data(v):
    v = (v or "").strip()
    if not v:
        return None
    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(v, fmt)
        except ValueError:
            continue
    return None


def _iw39_ordens(s):
    _drain(s)
    s.findById("wnd[0]/tbar[0]/okcd").Text = "/niw39"
    s.findById("wnd[0]").sendVKey(0)
    time.sleep(1.0)
    for idp in ("ctxtAUFNR-LOW", "ctxtAUFNR-HIGH", "ctxtGEWRK-LOW", "ctxtARBPL-LOW",
                "ctxtAUART-LOW", "ctxtAUART-HIGH", "ctxtSTAE1-LOW", "ctxtSTAE1-HIGH",
                "ctxtSTAI1-LOW"):
        try:
            s.findById("wnd[0]/usr/" + idp).Text = ""
        except Exception:
            pass
    try:
        s.findById("wnd[0]/usr/ctxtVARIANT").Text = VARIANTE  # layout determinístico
    except Exception:
        pass
    s.findById("wnd[0]/usr/ctxtIWERK-LOW").Text = "BR02"
    # "sem filtro de data" na prática: range amplo (datas VAZIAS abrem modal e travam)
    try:
        s.findById("wnd[0]/usr/ctxtDATUV").Text = "01.01.2010"
        s.findById("wnd[0]/usr/ctxtDATUB").Text = "31.12.2035"
    except Exception:
        pass
    _clip_set(AUART_BACKLOG)
    s.findById("wnd[0]/usr/ctxtAUART-LOW").setFocus()
    s.findById("wnd[0]/usr/btn%_AUART_%_APP_%-VALU_PUSH").press()
    time.sleep(0.6)
    s.findById("wnd[1]/tbar[0]/btn[24]").press()
    time.sleep(0.5)
    s.findById("wnd[1]/tbar[0]/btn[8]").press()
    time.sleep(0.5)
    for cb, v in {"DY_OFN": True, "DY_IAR": True, "DY_MAB": False, "DY_HIS": False}.items():
        try:
            s.findById("wnd[0]/usr/chk" + cb).Selected = v
        except Exception:
            pass
    _log("IW39: F8...")
    s.findById("wnd[0]").sendVKey(8)
    time.sleep(2.0)
    try:
        if s.findById("wnd[1]", False) is not None:
            s.findById("wnd[1]").sendVKey(0)
            time.sleep(0.5)
    except Exception:
        pass
    g = _find_grid(s)
    if g is None:
        _log("IW39: sem grid")
        return {}
    _log(f"IW39: grid {g.RowCount} linhas -> exportando xlsx")
    target = os.path.join(tempfile.gettempdir(), "backlog_iw39.xlsx")
    # PLTXT = denominação do local de instalação (texto). Best-effort: só vem se a variante
    # /BACKLOG do IW39 expuser a coluna. Confirmar o tech-name com o agente SAP (IM-17).
    recs = _export_and_read(s, g, ["AUFNR", "AUART", "KTEXT", "TPLNR", "PLTXT", "GSTRP"], target)
    od = {}
    for r in recs:
        a = (r.get("AUFNR") or "").strip()
        if a and a not in od:
            od[a] = r
    _log(f"IW39: {len(od)} ordens")
    return od


def _iw37_centro_op0010(s, ordens):
    centro = {}
    if not ordens:
        return centro
    _drain(s)
    s.findById("wnd[0]/tbar[0]/okcd").Text = "/niw37"
    s.findById("wnd[0]").sendVKey(0)
    time.sleep(1.0)
    for idp in ("ctxtAUFNR-LOW", "ctxtAUFNR-HIGH", "ctxtARBPL-LOW", "ctxtAUART-LOW",
                "ctxtDATUV", "ctxtDATUB"):
        try:
            s.findById("wnd[0]/usr/" + idp).Text = ""
        except Exception:
            pass
    try:
        s.findById("wnd[0]/usr/ctxtVARIANT").Text = VARIANTE  # /BACKLOG já traz ARBPL
    except Exception:
        pass
    _clip_set(list(ordens))
    s.findById("wnd[0]/usr/ctxtAUFNR-LOW").setFocus()
    s.findById("wnd[0]/usr/btn%_AUFNR_%_APP_%-VALU_PUSH").press()
    time.sleep(0.8)
    s.findById("wnd[1]/tbar[0]/btn[24]").press()
    time.sleep(0.8)
    s.findById("wnd[1]/tbar[0]/btn[8]").press()
    time.sleep(0.8)
    for cb in ("DY_AKT", "DY_PMONL"):
        try:
            s.findById("wnd[0]/usr/chk" + cb).Selected = True
        except Exception:
            pass
    try:
        s.findById("wnd[0]/usr/chkDY_HIS").Selected = False
    except Exception:
        pass
    _log("IW37: F8...")
    s.findById("wnd[0]").sendVKey(8)
    time.sleep(2.5)
    try:
        if s.findById("wnd[1]", False) is not None:
            s.findById("wnd[1]").sendVKey(0)
            time.sleep(0.5)
    except Exception:
        pass
    g = _find_grid(s)
    if g is None:
        _log("IW37: sem grid")
        return centro
    cols = [c for c in _coll(g.ColumnOrder) if c is not None]
    if "ARBPL" not in cols:
        # fallback defensivo: a variante /BACKLOG já deveria trazer ARBPL; só cai aqui
        # se a variante foi removida/alterada no SAP.
        _log("IW37: ARBPL ausente na variante -> fallback: expondo via layout...")
        try:
            s.findById("wnd[0]/tbar[1]/btn[32]").press()
            time.sleep(1.2)
            if s.findById("wnd[1]", False) is None:
                raise RuntimeError("dialog layout nao abriu")
            base = ("wnd[1]/usr/tabsG_TS_ALV/tabpALV_M_R1/"
                    "ssubSUB_DYN0510:SAPLSKBH:0620/")
            hidden = s.findById(base + "cntlCONTAINER1_LAYO/shellcont/shell")
            nset = hidden.RowCount
            if nset:
                hidden.selectedRows = f"0-{nset - 1}"
            s.findById(base + "btnAPP_WL_SING").press()
            time.sleep(0.8)
            s.findById("wnd[1]/tbar[0]/btn[0]").press()
            time.sleep(1.5)
            g = _find_grid(s)
            cols = [c for c in _coll(g.ColumnOrder) if c is not None]
        except Exception as ex:
            _log(f"IW37: expor ARBPL falhou: {ex}")
            _drain(s)
    if "ARBPL" not in cols:
        _log("IW37: ARBPL ausente -> centros ficam None")
        return centro
    _log(f"IW37: grid {g.RowCount} linhas -> exportando xlsx")
    target = os.path.join(tempfile.gettempdir(), "backlog_iw37.xlsx")
    recs = _export_and_read(s, g, ["AUFNR", "VORNR", "ARBPL"], target)
    for r in recs:
        a = (r.get("AUFNR") or "").strip()
        c = (r.get("ARBPL") or "").strip()
        v = (r.get("VORNR") or "").strip()
        if not a or not c:
            continue
        try:
            is_0010 = int(v) == 10
        except ValueError:
            is_0010 = (v == "0010")
        if is_0010:
            centro[a] = c
    _log(f"IW37: centro op0010 p/ {len(centro)} ordens")
    return centro


def coletar_ordens(session) -> list[dict]:
    """Coleta o backlog cru da BR02 (IW39 + IW37). READ-ONLY. Devolve list[dict].
    Não classifica. Completude > pureza. Contrato de saída: ver módulo."""
    s = session
    ordens = _iw39_ordens(s)
    centros = _iw37_centro_op0010(s, list(ordens))
    out = []
    for a in sorted(ordens, key=lambda x: int(x) if x.isdigit() else x):
        r = ordens[a]
        out.append({
            "ordem": a,
            "tipo": (r.get("AUART") or "").strip(),
            "descricao": (r.get("KTEXT") or "").strip(),
            "local_instalacao": (r.get("TPLNR") or "").strip(),
            "local_descricao": (r.get("PLTXT") or "").strip(),
            "data_inicio": _parse_data(r.get("GSTRP")),
            "centro_operacao_0010": centros.get(a),
        })
    try:
        s.findById("wnd[0]/tbar[0]/okcd").Text = "/n"
        s.findById("wnd[0]").sendVKey(0)
    except Exception:
        pass
    return out


def mock_ordens() -> list[dict]:
    """Amostra para teste E2E sem SAP (ctx.cfg.mock_sap=True). Cobre as 4 categorias,
    Mecânica/Elétrica/Não classificado e uma preventiva não-vencida (futura)."""
    return [
        # Corretivas (YPM1) — mecânica, elétrica, não classificado
        {"ordem": "5012522", "tipo": "YPM1", "descricao": "TROCA ROLAMENTO PRENSA 01",
         "local_instalacao": "BR02-0100-1600", "data_inicio": datetime(2026, 5, 13),
         "centro_operacao_0010": "Z10"},
        {"ordem": "5013391", "tipo": "YPM1", "descricao": "PAINEL ELETRICO LINHA 2",
         "local_instalacao": "BR02-0100-1300", "data_inicio": datetime(2026, 4, 11),
         "centro_operacao_0010": "Z20"},
        {"ordem": "5019639", "tipo": "YPM1", "descricao": "ORDEM NOVA SEM RECLASSIFICAR",
         "local_instalacao": "BR02-0700-0200", "data_inicio": datetime(2026, 6, 20),
         "centro_operacao_0010": "Z02"},
        # Preventivas (YPM2) — vencida (mecânica), preditiva (mecânica), FUTURA (não exibe)
        {"ordem": "5012523", "tipo": "YPM2", "descricao": "CALIBRACAO BALANCA ROD",
         "local_instalacao": "BR02-0100-1600", "data_inicio": datetime(2026, 2, 21),
         "centro_operacao_0010": "Z13"},
        {"ordem": "5010812", "tipo": "YPM2", "descricao": "PREVENTIVA CENTRAL EMG - ELE",
         "local_instalacao": "BR02-0700-0200", "data_inicio": datetime(2026, 5, 25),
         "centro_operacao_0010": "Z21"},
        {"ordem": "5099999", "tipo": "YPM2", "descricao": "PREVENTIVA FUTURA (NAO VENCIDA)",
         "local_instalacao": "BR02-0100-0300", "data_inicio": datetime(2026, 12, 15),
         "centro_operacao_0010": "Z10"},
        # Melhoria (YPM8) — sem disciplina
        {"ordem": "5019529", "tipo": "YPM8", "descricao": "(TEK)MELHORIA SINOPTICO IHM PRENSA",
         "local_instalacao": "BR02-0100-1600", "data_inicio": datetime(2026, 6, 1),
         "centro_operacao_0010": "Z10"},
        # Matrizes (YPM9) — sem disciplina
        {"ordem": "5019628", "tipo": "YPM9", "descricao": "MANUT MATRIZ ESTAMPO",
         "local_instalacao": "BR02-0100-1500", "data_inicio": datetime(2026, 5, 30),
         "centro_operacao_0010": "Z12"},
    ]
